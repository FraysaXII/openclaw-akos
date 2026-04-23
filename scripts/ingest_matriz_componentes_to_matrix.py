#!/usr/bin/env python3
"""One-off: Matriz componentes.xlsx 'components' sheet -> COMPONENT_SERVICE_MATRIX.csv.

Run from repo root: py scripts/ingest_matriz_componentes_to_matrix.py
Requires: openpyxl
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from akos.hlk_component_service_csv import COMPONENT_SERVICE_FIELDNAMES
from akos.io import REPO_ROOT

XLSX = REPO_ROOT / "Matriz componentes.xlsx"
OUT = REPO_ROOT / "docs" / "references" / "hlk" / "compliance" / "COMPONENT_SERVICE_MATRIX.csv"

VALID_ROLES = {
    "System Owner",
    "DevOPS",
    "CTO",
    "Data Architect",
    "AI Engineer",
    "Tech Lead",
    "CFO",
    "Business Controller",
    "PMO",
    "Compliance",
}


def norm_role(cell) -> str:
    if cell is None:
        return ""
    s = str(cell).strip()
    if s in VALID_ROLES:
        return s
    return ""


def kind_from_cat(cat: str | None, sub: str | None) -> str:
    c = (cat or "").lower()
    s = (sub or "").lower()
    if "hosting" in c or "domain" in s:
        return "infrastructure"
    if "finance" in c or "bank" in s:
        return "other"
    if "storage" in c and "api" in s:
        return "saas"
    if "api" in s or "api" in c:
        return "integration"
    if "data" in c:
        return "data_platform"
    return "saas"


def steward_from_cat(cat: str | None) -> str:
    c = (cat or "").lower()
    if "finance" in c:
        return "FINOPS"
    if "hosting" in c or "infrastructure" in c:
        return "DEVOPS"
    if "data" in c:
        return "DATAOPS"
    return "DEVOPS"


def api_exposure_from_row(sub: str | None, product: str | None) -> str:
    s = f"{sub or ''} {product or ''}".lower()
    if "postman" in s or "api" in s:
        return "internal"
    return "none"


def env_scope(env: str | None) -> str:
    e = (env or "").strip().lower()
    if e == "cloud":
        return "multi"
    return "multi"


def clean_url(u) -> str:
    if u is None or u == "":
        return ""
    s = str(u).strip()
    if s.lower() in ("none", "n/a"):
        return ""
    return s


def esc_csv(s: str) -> str:
    return s.replace("\n", " ").replace("\r", " ")


def main() -> int:
    if not XLSX.is_file():
        print("Missing", XLSX, file=sys.stderr)
        return 1

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "components" not in wb.sheetnames:
        print("No 'components' sheet", file=sys.stderr)
        return 1
    ws = wb["components"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(headers) if h}

    def g(row: tuple, key: str):
        i = idx.get(key)
        if i is None or i >= len(row):
            return None
        return row[i]

    rows_out: list[dict[str, str]] = []
    seen_display: set[str] = set()
    today = "2026-04-20"

    for rnum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        module = g(row, "module")
        product = g(row, "product") or module
        provider = (g(row, "provider") or "").strip() if g(row, "provider") else ""
        if not product and not module:
            continue
        name = str(product or module).strip()
        if not name:
            continue
        cid = f"comp_matriz_{rnum:05d}"
        base = f"{name} — {provider}" if provider else name
        disp_name = base
        if disp_name in seen_display:
            disp_name = f"{base} (sheet row {rnum})"
        seen_display.add(disp_name)

        cat = g(row, "category")
        sub = g(row, "subcategory")
        obj_class = g(row, "Object Class")
        desc = g(row, "Description")

        pr = norm_role(g(row, "System Owner")) or "System Owner"
        cdu = norm_role(g(row, "Chief Data User"))
        du = norm_role(g(row, "Data User"))
        sec = du or cdu

        notes_parts = [
            f"category={cat}",
            f"subcategory={sub}",
        ]
        if obj_class:
            notes_parts.append(f"object_class={obj_class}")
        if desc:
            notes_parts.append(f"description={esc_csv(str(desc)[:500])}")

        row_dict = {
            "component_id": cid,
            "component_name": disp_name[:500],
            "component_kind": kind_from_cat(
                str(cat) if cat is not None else None,
                str(sub) if sub is not None else None,
            ),
            "lifecycle_status": "active",
            "entity": "HLK Tech Lab",
            "area": "Tech",
            "primary_owner_role": pr,
            "steward_ops_domain": steward_from_cat(str(cat) if cat is not None else None),
            "secondary_owner_role": sec,
            "escalation_owner_role": "System Owner",
            "repo_slug": "",
            "github_url": "",
            "api_exposure": api_exposure_from_row(
                str(sub) if sub is not None else None,
                str(product) if product is not None else None,
            ),
            "api_spec_pointer": "",
            "integration_pattern": "n_a",
            "depends_on_component_ids": "",
            "parent_component_id": "",
            "primary_process_item_id": "",
            "related_process_item_ids": "",
            "topic_ids": "",
            "access_level_data": "3",
            "data_classification": "internal",
            "environment_scope": env_scope(str(g(row, "environment") or "")),
            "slo_tier": "standard",
            "runbook_link": clean_url(g(row, "hlk_documentation_url")),
            "doc_link": clean_url(g(row, "supplier_documentation_url")),
            "legal_hold": "",
            "retention_policy_ref": "",
            "last_verified_date": today,
            "source_row": "matriz_componentes_xlsx",
            "notes": "; ".join(str(p) for p in notes_parts if p),
        }
        rows_out.append(row_dict)

    wb.close()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COMPONENT_SERVICE_FIELDNAMES, lineterminator="\n")
        w.writeheader()
        w.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows to {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
