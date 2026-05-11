"""Tests for scripts/extract_engagement_pdfs.py."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "scripts"))

import extract_engagement_pdfs as eep  # noqa: E402


def test_redact_text_masks_emails_and_phones() -> None:
    text = "Contact: alice@example.com or +33 6 12 34 56 78"
    out = eep.redact_text(text, [])
    assert "alice@example.com" not in out
    assert "[REDACTED-EMAIL]" in out
    assert "[REDACTED-PHONE]" in out


def test_redact_text_applies_operator_tokens_case_insensitive() -> None:
    text = "Béatrice is the lead at AcmeCorp; beatrice and acmecorp both leak."
    out = eep.redact_text(text, ["Béatrice", "Beatrice", "AcmeCorp"])
    assert "Béatrice" not in out
    assert "beatrice" not in out.lower().replace("[redacted]", "")
    assert "AcmeCorp" not in out
    assert "acmecorp" not in out.lower().replace("[redacted]", "")
    assert out.count("[REDACTED]") >= 4


def test_redact_text_preserves_safe_content() -> None:
    text = "WeBuy is the procure-to-pay system used by an enterprise customer."
    out = eep.redact_text(text, ["AcmeCorp"])
    assert "WeBuy" in out
    assert "procure-to-pay" in out


def test_safe_stem_normalises_spaces_and_accents() -> None:
    assert eep.safe_stem(Path("Mode opératoire - Process de passage.pdf")) == "mode_operatoire_-_process_de_passage"


def test_extraction_config_engagement_path(tmp_path: Path) -> None:
    cfg = eep.ExtractionConfig(
        mode="engagement",
        inputs=[],
        slug="2026-05-10-test-engagement",
        out_root_override=tmp_path,
    )
    assert cfg.out_root() == tmp_path


def test_extraction_config_engagement_requires_slug() -> None:
    cfg = eep.ExtractionConfig(mode="engagement", inputs=[], slug=None)
    with pytest.raises(ValueError):
        cfg.out_root()


def test_extraction_config_inspiration_default_outside_repo(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("HOLISTIKA_INSPIRATION_CACHE", str(tmp_path / "inspiration"))
    cfg = eep.ExtractionConfig(mode="inspiration", inputs=[], slug=None)
    assert cfg.out_root() == tmp_path / "inspiration"


def test_extract_passthrough_md_roundtrips(tmp_path: Path) -> None:
    src = tmp_path / "transcript.md"
    src.write_text("# transcript\nbody line", encoding="utf-8")
    body = eep.extract_one(src)
    assert "transcript" in body and "body line" in body


def test_extract_json_roundtrips(tmp_path: Path) -> None:
    src = tmp_path / "data.json"
    src.write_text(json.dumps({"a": 1, "b": "alice@example.com"}), encoding="utf-8")
    body = eep.extract_one(src)
    assert "alice@example.com" in body
    redacted = eep.redact_text(body, [])
    assert "alice@example.com" not in redacted


def test_extract_xlsx_roundtrips_basic_workbook(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    import openpyxl

    src = tmp_path / "tiny.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing"
    ws["A1"] = "role"
    ws["B1"] = "rate_eur_h"
    ws["A2"] = "Senior Researcher"
    ws["B2"] = 80
    wb.save(src)
    body = eep.extract_one(src)
    assert "sheet Pricing" in body
    assert "Senior Researcher" in body
    assert "80" in body


def test_extract_one_rejects_unsupported_suffix(tmp_path: Path) -> None:
    src = tmp_path / "foo.bin"
    src.write_bytes(b"\x00\x01")
    with pytest.raises(ValueError):
        eep.extract_one(src)


def test_run_extract_handles_missing_input(tmp_path: Path) -> None:
    cfg = eep.ExtractionConfig(
        mode="engagement",
        inputs=[tmp_path / "does-not-exist.pdf"],
        slug="2026-05-10-test",
        out_root_override=tmp_path / "out",
    )
    failures = eep.run_extract(cfg)
    assert failures == 1


def test_run_extract_writes_redacted_output(tmp_path: Path) -> None:
    src = tmp_path / "note.md"
    src.write_text("Hello Béatrice at alice@example.com.", encoding="utf-8")
    out_root = tmp_path / "out"
    cfg = eep.ExtractionConfig(
        mode="engagement",
        inputs=[src],
        slug="2026-05-10-test",
        redact_tokens=["Béatrice"],
        out_root_override=out_root,
    )
    failures = eep.run_extract(cfg)
    assert failures == 0
    written = out_root / "note.txt"
    assert written.exists()
    body = written.read_text(encoding="utf-8")
    assert "Béatrice" not in body
    assert "alice@example.com" not in body
